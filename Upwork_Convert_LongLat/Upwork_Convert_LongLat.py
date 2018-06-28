import pprint, json
import urllib, urllib.request
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename='Input.xlsx')
wb.guess_types = True
ws = wb.worksheets[0]

Longitude = []
Latitude = []

for row in ws.iter_rows('A{}:A{}'.format(ws.min_row,ws.max_row)):
    for cell in row:
        Longitude.append(cell.value)
print("Reading Longitude Done") # Setup Reading CSV"""

for row in ws.iter_rows('B{}:B{}'.format(ws.min_row,ws.max_row)):
    for cell in row:
        Latitude.append(cell.value)
print("Reading Latitude Done")

f = open('Output.txt','a')
for var in range(len(Latitude)):
    if(var == 0):
        temp = 1
        url = ("https://maps.googleapis.com/maps/api/geocode/json?latlng={},{}&key=AIzaSyDeyQLXT5Fo6WYVvzQ9XSIbaXURTmbydV8".format(Longitude[temp],Latitude[temp]))
        response = urllib.request.urlopen(url)
        data = json.loads(response.read())
        if(str(data['status']) == 'ZERO_RESULTS'):
            f.write('ZERO_RESULTS' + '\n')
        else:
            f.write(str(data['results'][0]['formatted_address']) + '\n')
            continue
    url = ("https://maps.googleapis.com/maps/api/geocode/json?latlng={},{}&key=AIzaSyDeyQLXT5Fo6WYVvzQ9XSIbaXURTmbydV8".format(Longitude[var],Latitude[var]))
    response = urllib.request.urlopen(url)
    data = json.loads(response.read())
    if(str(data['status']) == 'ZERO_RESULTS'):
        f.write('ZERO_RESULTS' + '\n')
    else:
        f.write(str(data['results'][0]['formatted_address']) + '\n')
f.close()

